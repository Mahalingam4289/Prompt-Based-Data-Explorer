import warnings
warnings.filterwarnings("ignore")

import re
import io
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
import seaborn as sns
from scipy import stats
from sklearn.preprocessing import LabelEncoder, StandardScaler, MinMaxScaler
from sklearn.impute import KNNImputer
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.metrics import classification_report, confusion_matrix, roc_auc_score
from sklearn.pipeline import Pipeline
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, KeepTogether, PageBreak
)
from reportlab.platypus import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import streamlit as st

st.set_page_config(
    page_title="EduQuery Explorer",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=JetBrains+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

.stApp { background: #0B0F1A; color: #EEF2FF; }

[data-testid="stSidebar"] {
    background: #111827;
    border-right: 1px solid #243050;
}
[data-testid="stSidebar"] * { color: #EEF2FF !important; }

.header-banner {
    background: linear-gradient(135deg, #0B1020 0%, #111827 100%);
    border: 1px solid #243050;
    border-radius: 12px;
    padding: 18px 24px;
    margin-bottom: 20px;
    display: flex;
    align-items: center;
    gap: 14px;
}
.header-title { font-size: 22px; font-weight: 700; color: #EEF2FF; margin: 0; }
.header-sub   { font-size: 12px; color: #6B7DB3; font-family: 'JetBrains Mono', monospace; margin-top: 2px; }

.pipeline-wrap {
    display: flex; align-items: center; gap: 4px;
    background: #111827; border: 1px solid #243050;
    border-radius: 10px; padding: 10px 16px;
    margin-bottom: 20px; overflow-x: auto;
}
.pipe-step {
    display: flex; align-items: center; gap: 5px;
    padding: 4px 10px; border-radius: 6px;
    border: 1px solid #243050;
    font-family: 'JetBrains Mono', monospace;
    font-size: 10px; color: #6B7DB3;
    white-space: nowrap; transition: all .3s;
}
.pipe-step.done   { border-color:rgba(16,185,129,.4); background:rgba(16,185,129,.08); color:#10B981; }
.pipe-step.active { border-color:rgba(245,158,11,.5); background:rgba(245,158,11,.1);  color:#FCD34D; }
.pipe-arrow { color: #3D4F7C; font-size: 12px; }

.metric-row { display: flex; gap: 12px; margin-bottom: 18px; flex-wrap: wrap; }
.metric-card {
    flex: 1; min-width: 130px;
    background: #1A2235; border: 1px solid #243050;
    border-radius: 10px; padding: 14px 16px;
}
.metric-label { font-size: 10px; color: #6B7DB3; text-transform: uppercase; letter-spacing: .08em; margin-bottom: 4px; }
.metric-value { font-size: 22px; font-weight: 700; color: #F59E0B; font-family: 'JetBrains Mono', monospace; }
.metric-sub   { font-size: 10px; color: #6B7DB3; margin-top: 2px; }

.msg-user {
    background: linear-gradient(135deg,rgba(245,158,11,.14),rgba(245,158,11,.07));
    border: 1px solid rgba(245,158,11,.35);
    border-radius: 16px 16px 4px 16px;
    padding: 12px 16px; margin: 8px 0; margin-left: 60px;
    font-size: 13px; line-height: 1.6;
}
.msg-assistant {
    background: #1A2235; border: 1px solid #243050;
    border-radius: 4px 16px 16px 16px;
    padding: 14px 16px; margin: 8px 0; margin-right: 20px;
}
.msg-meta { font-size: 10px; color: #3D4F7C; font-family: 'JetBrains Mono', monospace; margin-bottom: 6px; }
.msg-title { font-size: 16px; font-weight: 700; color: #EEF2FF; margin-bottom: 10px; }

.nlp-trace {
    background: rgba(16,185,129,.04); border: 1px solid rgba(16,185,129,.2);
    border-radius: 8px; padding: 10px 14px; margin: 10px 0;
    font-family: 'JetBrains Mono', monospace; font-size: 11px;
}
.nlp-title { font-size: 9px; color: #10B981; font-weight: 600; letter-spacing: .1em; text-transform: uppercase; margin-bottom: 7px; }
.nlp-tag-intent { display:inline-block; padding:2px 8px; border-radius:10px; font-size:10px; background:rgba(245,158,11,.1); border:1px solid rgba(245,158,11,.3); color:#FCD34D; margin:2px; }
.nlp-tag-field  { display:inline-block; padding:2px 8px; border-radius:10px; font-size:10px; background:rgba(59,130,246,.1);  border:1px solid rgba(59,130,246,.3);  color:#60A5FA; margin:2px; }
.nlp-tag-group  { display:inline-block; padding:2px 8px; border-radius:10px; font-size:10px; background:rgba(139,92,246,.1);  border:1px solid rgba(139,92,246,.3);  color:#C084FC; margin:2px; }
.nlp-tag-filter { display:inline-block; padding:2px 8px; border-radius:10px; font-size:10px; background:rgba(16,185,129,.1);  border:1px solid rgba(16,185,129,.3);  color:#34D399; margin:2px; }

.insight-box {
    background: rgba(16,185,129,.06); border: 1px solid rgba(16,185,129,.25);
    border-radius: 8px; padding: 12px 14px; margin-top: 10px;
}
.insight-title { font-size: 9px; color: #10B981; font-family: 'JetBrains Mono', monospace;
                 letter-spacing:.1em; text-transform:uppercase; margin-bottom: 5px; }
.insight-text  { font-size: 13px; color: #EEF2FF; line-height: 1.6; }

/* ML Risk box */
.ml-box {
    background: rgba(139,92,246,.06); border: 1px solid rgba(139,92,246,.25);
    border-radius: 8px; padding: 12px 14px; margin-top: 10px;
}
.ml-title { font-size: 9px; color: #C084FC; font-family: 'JetBrains Mono', monospace;
            letter-spacing:.1em; text-transform:uppercase; margin-bottom: 5px; }

.prep-item {
    display: flex; gap: 8px; align-items: flex-start;
    padding: 5px 0; border-bottom: 1px solid #1A2235;
}
.prep-check {
    width: 16px; height: 16px; border-radius: 4px; flex-shrink: 0; margin-top: 1px;
    background: rgba(16,185,129,.15); border: 1px solid rgba(16,185,129,.35);
    display: flex; align-items: center; justify-content: center;
    font-size: 9px; color: #10B981;
}
.prep-name   { font-size: 11px; font-weight: 600; color: #EEF2FF; }
.prep-detail { font-size: 10px; color: #6B7DB3; margin-top: 1px; }

.stButton > button {
    background: transparent !important;
    border: 1px solid #243050 !important;
    color: #6B7DB3 !important;
    border-radius: 20px !important;
    font-size: 11px !important;
    padding: 4px 12px !important;
    transition: all .15s !important;
}
.stButton > button:hover {
    border-color: #F59E0B !important;
    color: #FCD34D !important;
}

.stTextInput > div > div > input,
.stTextArea > div > div > textarea {
    background: #1A2235 !important; color: #EEF2FF !important;
    border: 1px solid #243050 !important; border-radius: 10px !important;
    font-family: 'DM Sans', sans-serif !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: rgba(245,158,11,.6) !important;
    box-shadow: 0 0 0 2px rgba(245,158,11,.1) !important;
}

[data-testid="stFileUploader"] {
    background: #1A2235; border: 2px dashed #243050;
    border-radius: 10px; padding: 8px;
}

hr { border-color: #243050; }
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-thumb { background: #243050; border-radius: 4px; }

.badge-green { display:inline-block; padding:3px 10px; border-radius:20px; font-size:10px; 
               color:#10B981; border:1px solid rgba(16,185,129,.35); background:rgba(16,185,129,.08); 
               font-family:'JetBrains Mono',monospace; }
.badge-amber { display:inline-block; padding:3px 10px; border-radius:20px; font-size:10px; 
               color:#FCD34D; border:1px solid rgba(245,158,11,.35); background:rgba(245,158,11,.08); 
               font-family:'JetBrains Mono',monospace; }
.badge-purple { display:inline-block; padding:3px 10px; border-radius:20px; font-size:10px; 
               color:#C084FC; border:1px solid rgba(139,92,246,.35); background:rgba(139,92,246,.08); 
               font-family:'JetBrains Mono',monospace; }
</style>
""", unsafe_allow_html=True)

EXPECTED_COLS = [
    "Stu_ID", "Name", "Gender", "DOB", "State", "Department", "Program",
    "Enrollment_Status", "Current_Course", "Credits_Earned", "GPA",
    "Attendance_Pct", "Financial_Aid_Type", "Parent_Survey",
    "School_Satisfaction", "Absent_Days", "School_Level",
]

CHART_COLORS = [
    "#F59E0B", "#3B82F6", "#10B981", "#8B5CF6", "#EF4444",
    "#06B6D4", "#F97316", "#EC4899", "#84CC16", "#A78BFA",
    "#FCD34D", "#60A5FA", "#34D399", "#C084FC", "#FB7185",
]

QUICK_PROMPTS = [
    "Show average GPA by department",
    "Compare attendance across gender",
    "Top 5 departments by GPA",
    "Distribution of GPA scores",
    "How many students by school level",
    "Count students with above-7 absent days",
    "Correlation between GPA and attendance",
    "Show at-risk students by department",
    "Filter students with low attendance",
    "Distribution of financial aid types",
    "Rank departments by average attendance",
    "Summary of the dataset",
]

PIPELINE_STEPS = [
    ("⬆", "Dataset Upload"),
    ("⚙", "Preprocessing"),
    ("✎", "Prompt Input"),
    ("◈", "NLP Processing"),
    ("⊕", "Query Generation"),
    ("◎", "Data Analysis"),
    ("◉", "Visualization"),
    ("★", "Result & Insights"),
]


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — PREPROCESSING ENGINE  (unchanged from original)
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def preprocess_dataframe(file_bytes, filename):
    log = []
    if filename.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes))
    else:
        try:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Combined Dataset", header=2)
        except Exception:
            df = pd.read_excel(io.BytesIO(file_bytes), header=0)

    df.columns = [
        str(c).strip().replace(" ", "_").replace("%", "Pct")
        for c in df.columns
    ]
    df.columns = EXPECTED_COLS[: len(df.columns)]
    df = df[df["Stu_ID"].astype(str).str.strip().ne("")]
    df = df[df["Stu_ID"].astype(str).str.strip().ne("nan")]
    log.append({"step": "Column Normalization",
                "detail": f"{len(df.columns)} columns mapped → {len(df):,} rows"})

    miss_cols = [c for c in df.columns if df[c].isnull().sum() > 0]
    log.append({"step": "Missing Value Analysis",
                "detail": f"{len(miss_cols)} columns with nulls detected"})

    df["DOB"]            = pd.to_datetime(df["DOB"], errors="coerce")
    df["GPA"]            = pd.to_numeric(df["GPA"],            errors="coerce")
    df["Attendance_Pct"] = pd.to_numeric(df["Attendance_Pct"], errors="coerce")
    df["Credits_Earned"] = pd.to_numeric(df["Credits_Earned"], errors="coerce")
    log.append({"step": "Data Type Correction",
                "detail": "GPA, Attendance_Pct, Credits_Earned → float64; DOB → datetime"})

    num_cols = ["GPA", "Attendance_Pct", "Credits_Earned"]
    num_na   = [c for c in num_cols if df[c].isnull().any()]
    if num_na:
        imp = KNNImputer(n_neighbors=5)
        df[num_cols] = imp.fit_transform(df[num_cols].fillna(df[num_cols].median()))
        df["GPA"] = df["GPA"].clip(0.0, 4.0).round(2)

    cat_na = [c for c in df.select_dtypes("object").columns if df[c].isnull().any()]
    for c in cat_na:
        df[c].fillna(df[c].mode()[0], inplace=True)
    log.append({"step": "Missing Value Imputation",
                "detail": f"KNN imputer on {num_cols}; mode fill on categoricals"})

    n_before = len(df)
    df.drop_duplicates(inplace=True)
    df.reset_index(drop=True, inplace=True)
    log.append({"step": "Duplicate Removal",
                "detail": f"Removed {n_before - len(df)} duplicates → {len(df):,} rows remain"})

    for col in ["GPA", "Attendance_Pct", "Credits_Earned"]:
        vals = df[col].dropna()
        Q1, Q3 = vals.quantile(0.25), vals.quantile(0.75)
        IQR = Q3 - Q1
        lo, hi = Q1 - 1.5 * IQR, Q3 + 1.5 * IQR
        df[col] = df[col].clip(lower=lo, upper=hi)
    log.append({"step": "Outlier Treatment",
                "detail": "IQR Winsorization on GPA, Attendance_Pct, Credits_Earned"})

    def perf_band(g):
        if g >= 3.7: return "Distinction"
        if g >= 3.0: return "Merit"
        if g >= 2.0: return "Pass"
        return "At-Risk"

    df["Age"] = df["DOB"].apply(
        lambda d: int((pd.Timestamp("2024-06-01") - d).days // 365)
        if pd.notnull(d) else np.nan
    )
    df["Performance_Band"]  = df["GPA"].apply(perf_band)
    df["Low_Attendance"]    = (df["Attendance_Pct"] < 75).astype(int)
    df["Credits_Progress"]  = (df["Credits_Earned"] / 120.0).clip(0, 1).round(3)
    df["Engagement_Score"]  = (
        (df["GPA"] / 4.0) * 0.5
        + (df["Attendance_Pct"] / 100.0) * 0.3
        + df["Parent_Survey"].eq("Yes").astype(float) * 0.1
        + df["School_Satisfaction"].eq("Good").astype(float) * 0.1
    ).round(3)
    df["Risk_Index"] = (
        (1 - df["GPA"] / 4.0) * 0.4
        + (1 - df["Attendance_Pct"] / 100.0) * 0.3
        + df["Absent_Days"].eq("Above-7").astype(float) * 0.2
        + df["Parent_Survey"].ne("Yes").astype(float) * 0.05
        + df["School_Satisfaction"].ne("Good").astype(float) * 0.05
    ).round(3)

    def credits_to_year(c):
        c = float(c) if pd.notna(c) else 0.0
        if c < 30:  return "Freshman"
        if c < 60:  return "Sophomore"
        if c < 90:  return "Junior"
        return "Senior"

    df["Academic_Year"] = df["Credits_Earned"].apply(credits_to_year)
    log.append({"step": "Feature Engineering",
                "detail": "Added: Performance_Band, Engagement_Score, Risk_Index, Academic_Year, Low_Attendance"})

    le_cols = ["Gender", "Enrollment_Status", "School_Satisfaction",
               "Absent_Days", "School_Level", "Performance_Band"]
    for col in le_cols:
        if col in df.columns:
            le = LabelEncoder()
            df[f"{col}_LE"] = le.fit_transform(df[col].astype(str))
    log.append({"step": "Label Encoding",
                "detail": f"Encoded {len(le_cols)} categorical columns"})

    log.append({"step": "✅ Preprocessing Complete",
                "detail": f"Final dataset: {len(df):,} rows × {len(df.columns)} columns"})
    return df, log


# ══════════════════════════════════════════════════════════════════════════════
# NEW FEATURE 1 — PREDICTIVE RISK ML MODEL
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def train_risk_model(df_bytes):
    """
    Train a Random Forest classifier to predict At-Risk students.
    Returns model metrics + per-student risk probability DataFrame.
    Uses cached version keyed by df content hash.
    """
    import hashlib
    df = pd.read_json(io.StringIO(df_bytes))

    feature_cols = ["GPA", "Attendance_Pct", "Credits_Earned",
                    "Engagement_Score", "Low_Attendance",
                    "Gender_LE", "Enrollment_Status_LE",
                    "School_Satisfaction_LE", "Absent_Days_LE", "School_Level_LE"]
    feature_cols = [c for c in feature_cols if c in df.columns]

    df["Target"] = (df["Performance_Band"] == "At-Risk").astype(int)
    X = df[feature_cols].fillna(0)
    y = df["Target"]

    if y.sum() < 5 or (y == 0).sum() < 5:
        return None, "Insufficient at-risk samples for ML training."

    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.25, random_state=42, stratify=y
    )

    model = Pipeline([
        ("scaler", StandardScaler()),
        ("clf", RandomForestClassifier(
            n_estimators=150, max_depth=8,
            class_weight="balanced", random_state=42, n_jobs=-1
        ))
    ])
    model.fit(X_train, y_train)

    y_prob  = model.predict_proba(X_test)[:, 1]
    y_pred  = model.predict(X_test)
    auc     = round(roc_auc_score(y_test, y_prob), 3)
    cv_scores = cross_val_score(model, X, y, cv=5, scoring="roc_auc")

    # Predict on full dataset
    all_prob = model.predict_proba(X)[:, 1]
    risk_df  = df[["Stu_ID", "Name", "Department", "GPA",
                   "Attendance_Pct", "Performance_Band"]].copy()
    risk_df["ML_Risk_Prob"]  = (all_prob * 100).round(1)
    risk_df["ML_Risk_Label"] = pd.cut(
        all_prob,
        bins=[0, 0.3, 0.6, 1.0],
        labels=["Low Risk", "Moderate Risk", "High Risk"]
    )

    # Feature importances
    rf = model.named_steps["clf"]
    fi = pd.DataFrame({
        "Feature": feature_cols,
        "Importance": rf.feature_importances_
    }).sort_values("Importance", ascending=False)

    # Confusion matrix
    cm = confusion_matrix(y_test, y_pred)

    metrics = {
        "auc":       auc,
        "cv_mean":   round(cv_scores.mean(), 3),
        "cv_std":    round(cv_scores.std(), 3),
        "n_train":   len(X_train),
        "n_test":    len(X_test),
        "n_features": len(feature_cols),
        "feature_importances": fi,
        "confusion_matrix": cm,
        "report": classification_report(y_test, y_pred, output_dict=True),
    }
    return risk_df.sort_values("ML_Risk_Prob", ascending=False), metrics


def render_ml_risk_panel(df):
    """Streamlit panel for ML risk model — shown in a dedicated tab."""
    st.markdown("""
    <div class="ml-box">
      <div class="ml-title">🤖 ML Module — Random Forest Risk Predictor</div>
      <div style="font-size:12px;color:#EEF2FF;line-height:1.6">
        Trains a <strong>Random Forest Classifier</strong> on 10 engineered features 
        to predict student dropout risk probability. Evaluated via 5-fold cross-validated AUC.
      </div>
    </div>
    """, unsafe_allow_html=True)

    df_json = df.to_json()

    with st.spinner("Training Random Forest model…"):
        risk_df, metrics = train_risk_model(df_json)

    if risk_df is None:
        st.error(metrics)
        return

    # ── Metric cards ──
    m = metrics
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("ROC-AUC Score", m["auc"], help="Higher = better discrimination")
    c2.metric("CV Mean AUC", m["cv_mean"], f"±{m['cv_std']}")
    c3.metric("Training Samples", m["n_train"])
    c4.metric("Test Samples", m["n_test"])

    # ── Feature importances chart ──
    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("**Feature Importances**")
        fi = m["feature_importances"]
        fig, ax = plt.subplots(figsize=(6, 4))
        fig.patch.set_facecolor("#111827")
        ax.set_facecolor("#1A2235")
        colors_fi = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(fi))]
        ax.barh(fi["Feature"][::-1], fi["Importance"][::-1],
                color=colors_fi[::-1], edgecolor="none", height=0.6)
        ax.set_xlabel("Importance", color="#6B7DB3", fontsize=9)
        ax.tick_params(colors="#EEF2FF", labelsize=8)
        for spine in ax.spines.values():
            spine.set_color("#243050")
        ax.xaxis.grid(True, color="#243050", linestyle="--", alpha=0.5)
        fig.tight_layout(pad=1)
        st.pyplot(fig, use_container_width=True)
        plt.close(fig)

    with col_b:
        st.markdown("**Confusion Matrix (Test Set)**")
        cm_arr = m["confusion_matrix"]
        fig, ax = plt.subplots(figsize=(4, 3.5))
        fig.patch.set_facecolor("#111827")
        ax.set_facecolor("#1A2235")
        im = ax.imshow(cm_arr, cmap="YlOrRd", aspect="auto")
        ax.set_xticks([0, 1]); ax.set_yticks([0, 1])
        ax.set_xticklabels(["Not At-Risk", "At-Risk"], color="#EEF2FF", fontsize=9)
        ax.set_yticklabels(["Not At-Risk", "At-Risk"], color="#EEF2FF", fontsize=9)
        ax.set_xlabel("Predicted", color="#6B7DB3", fontsize=9)
        ax.set_ylabel("Actual", color="#6B7DB3", fontsize=9)
        for i in range(2):
            for j in range(2):
                ax.text(j, i, str(cm_arr[i, j]), ha="center", va="center",
                        color="#0B0F1A", fontsize=14, fontweight="bold")
        fig.tight_layout(pad=1)
        st.pyplot(fig, use_container_width=True)
        plt.close(fig)

    # ── Risk probability distribution ──
    st.markdown("**ML Risk Probability Distribution**")
    fig, ax = plt.subplots(figsize=(10, 3))
    fig.patch.set_facecolor("#111827")
    ax.set_facecolor("#1A2235")
    probs = risk_df["ML_Risk_Prob"].values
    high = probs[probs >= 60]
    mod  = probs[(probs >= 30) & (probs < 60)]
    low  = probs[probs < 30]
    bins = np.linspace(0, 100, 30)
    ax.hist(low,  bins=bins, color="#10B981", alpha=0.85, label="Low Risk",      edgecolor="none")
    ax.hist(mod,  bins=bins, color="#F59E0B", alpha=0.85, label="Moderate Risk", edgecolor="none")
    ax.hist(high, bins=bins, color="#EF4444", alpha=0.85, label="High Risk",     edgecolor="none")
    ax.set_xlabel("Risk Probability (%)", color="#6B7DB3", fontsize=9)
    ax.set_ylabel("Student Count", color="#6B7DB3", fontsize=9)
    ax.tick_params(colors="#EEF2FF", labelsize=8)
    for spine in ax.spines.values(): spine.set_color("#243050")
    ax.legend(fontsize=9, frameon=False, labelcolor="#EEF2FF")
    fig.tight_layout(pad=1)
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)

    # ── High Risk student table ──
    st.markdown("**🔴 Top High-Risk Students (ML Predicted)**")
    high_risk = risk_df[risk_df["ML_Risk_Label"] == "High Risk"].head(20)
    st.dataframe(high_risk, use_container_width=True, hide_index=True)

    # ── Risk by department ──
    st.markdown("**Risk Distribution by Department**")
    dept_risk = (
        risk_df.groupby("Department")["ML_Risk_Prob"]
        .agg(["mean", "count"])
        .reset_index()
        .rename(columns={"mean": "Avg Risk %", "count": "Students"})
        .sort_values("Avg Risk %", ascending=False)
    )
    fig, ax = plt.subplots(figsize=(10, 4))
    fig.patch.set_facecolor("#111827")
    ax.set_facecolor("#1A2235")
    colors_dept = [
        "#EF4444" if v >= 50 else "#F59E0B" if v >= 30 else "#10B981"
        for v in dept_risk["Avg Risk %"]
    ]
    bars = ax.barh(dept_risk["Department"][::-1], dept_risk["Avg Risk %"][::-1],
                   color=colors_dept[::-1], edgecolor="none", height=0.6)
    ax.set_xlabel("Avg ML Risk Probability (%)", color="#6B7DB3", fontsize=9)
    ax.tick_params(colors="#EEF2FF", labelsize=8)
    for spine in ax.spines.values(): spine.set_color("#243050")
    ax.xaxis.grid(True, color="#243050", linestyle="--", alpha=0.5)
    fig.tight_layout(pad=1)
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)

    # ── Export ML results ──
    csv_buf = io.StringIO()
    risk_df.to_csv(csv_buf, index=False)
    st.download_button(
        "⬇ Download ML Risk Predictions (CSV)",
        csv_buf.getvalue(),
        file_name="ml_risk_predictions.csv",
        mime="text/csv",
    )


# ══════════════════════════════════════════════════════════════════════════════
# NEW FEATURE 2 — ADDITIONAL CHART TYPES
# ══════════════════════════════════════════════════════════════════════════════
def make_stacked_bar(df, group_col, stack_col, title=""):
    """Stacked bar — e.g. Performance_Band stacked by Department"""
    if group_col not in df.columns or stack_col not in df.columns:
        return None
    pivot = (
        df.groupby([group_col, stack_col])
        .size()
        .unstack(fill_value=0)
    )
    fig, ax = plt.subplots(figsize=(9, max(4, len(pivot) * 0.55 + 1)))
    fig.patch.set_facecolor("#111827")
    ax.set_facecolor("#1A2235")
    bottom = np.zeros(len(pivot))
    for i, col in enumerate(pivot.columns):
        vals = pivot[col].values.astype(float)
        ax.barh(range(len(pivot)), vals, left=bottom,
                color=CHART_COLORS[i % len(CHART_COLORS)],
                label=str(col), edgecolor="none", height=0.65)
        bottom += vals
    ax.set_yticks(range(len(pivot)))
    ax.set_yticklabels(pivot.index.tolist(), color="#EEF2FF", fontsize=9)
    ax.set_xlabel("Count", color="#6B7DB3", fontsize=9)
    ax.tick_params(colors="#6B7DB3", labelsize=9)
    for spine in ax.spines.values(): spine.set_color("#243050")
    ax.xaxis.grid(True, color="#243050", linestyle="--", alpha=0.5)
    ax.invert_yaxis()
    ax.legend(fontsize=8, frameon=False, labelcolor="#EEF2FF",
              loc="lower right", ncol=2)
    if title:
        ax.set_title(title, color="#EEF2FF", fontsize=11, pad=8)
    fig.tight_layout(pad=1.2)
    return fig


def make_scatter(df, x_col, y_col, color_col=None, title=""):
    """Scatter plot with optional color grouping"""
    if x_col not in df.columns or y_col not in df.columns:
        return None
    x = pd.to_numeric(df[x_col], errors="coerce")
    y = pd.to_numeric(df[y_col], errors="coerce")
    mask = x.notna() & y.notna()
    x, y = x[mask], y[mask]

    fig, ax = plt.subplots(figsize=(8, 5))
    fig.patch.set_facecolor("#111827")
    ax.set_facecolor("#1A2235")

    if color_col and color_col in df.columns:
        groups = df[color_col][mask].unique()
        for i, g in enumerate(groups):
            gm = df[color_col][mask] == g
            ax.scatter(x[gm], y[gm], color=CHART_COLORS[i % len(CHART_COLORS)],
                       alpha=0.6, s=18, label=str(g), edgecolors="none")
        ax.legend(fontsize=8, frameon=False, labelcolor="#EEF2FF")
    else:
        ax.scatter(x, y, color="#F59E0B", alpha=0.5, s=18, edgecolors="none")

    # Trend line
    if len(x) > 2:
        m_val, b_val, *_ = stats.linregress(x, y)
        x_line = np.linspace(x.min(), x.max(), 100)
        ax.plot(x_line, m_val * x_line + b_val,
                color="#EF4444", linewidth=1.5, linestyle="--", alpha=0.8)

    ax.set_xlabel(x_col.replace("_", " "), color="#6B7DB3", fontsize=9)
    ax.set_ylabel(y_col.replace("_", " "), color="#6B7DB3", fontsize=9)
    ax.tick_params(colors="#6B7DB3", labelsize=8)
    for spine in ax.spines.values(): spine.set_color("#243050")
    ax.grid(True, color="#243050", linestyle="--", alpha=0.4)
    if title:
        ax.set_title(title, color="#EEF2FF", fontsize=11, pad=8)
    fig.tight_layout(pad=1.2)
    return fig


def make_heatmap(df, row_col, col_col, val_col, title=""):
    """Heatmap of mean(val_col) grouped by row_col × col_col"""
    if not all(c in df.columns for c in [row_col, col_col, val_col]):
        return None
    pivot = (
        df.groupby([row_col, col_col])[val_col]
        .mean()
        .unstack(fill_value=0)
        .round(2)
    )
    fig, ax = plt.subplots(figsize=(max(6, len(pivot.columns) * 1.2),
                                    max(4, len(pivot) * 0.6 + 1)))
    fig.patch.set_facecolor("#111827")
    ax.set_facecolor("#1A2235")
    im = ax.imshow(pivot.values, cmap="YlOrRd", aspect="auto")
    ax.set_xticks(range(len(pivot.columns)))
    ax.set_yticks(range(len(pivot.index)))
    ax.set_xticklabels(pivot.columns.tolist(), color="#EEF2FF", fontsize=8,
                        rotation=30, ha="right")
    ax.set_yticklabels(pivot.index.tolist(), color="#EEF2FF", fontsize=8)
    for i in range(len(pivot.index)):
        for j in range(len(pivot.columns)):
            ax.text(j, i, f"{pivot.values[i, j]:.2f}",
                    ha="center", va="center", color="#0B0F1A", fontsize=8, fontweight="bold")
    cbar = plt.colorbar(im, ax=ax)
    cbar.ax.yaxis.set_tick_params(color="#6B7DB3")
    plt.setp(cbar.ax.yaxis.get_ticklabels(), color="#6B7DB3", fontsize=8)
    if title:
        ax.set_title(title, color="#EEF2FF", fontsize=11, pad=8)
    fig.tight_layout(pad=1.2)
    return fig


def make_box_plot(df, group_col, val_col, title=""):
    """Box plot — distribution of val_col across groups"""
    if group_col not in df.columns or val_col not in df.columns:
        return None
    groups = df[group_col].unique()
    data   = [pd.to_numeric(df[df[group_col] == g][val_col], errors="coerce").dropna()
              for g in groups]
    data   = [(g, d) for g, d in zip(groups, data) if len(d) > 0]
    if not data:
        return None
    groups, data = zip(*data)

    fig, ax = plt.subplots(figsize=(max(6, len(groups) * 0.9 + 1), 5))
    fig.patch.set_facecolor("#111827")
    ax.set_facecolor("#1A2235")
    bp = ax.boxplot(data, patch_artist=True, notch=False,
                    medianprops=dict(color="#0B0F1A", linewidth=2),
                    whiskerprops=dict(color="#6B7DB3"),
                    capprops=dict(color="#6B7DB3"),
                    flierprops=dict(marker="o", color="#EF4444", alpha=0.4, markersize=3))
    for patch, color in zip(bp["boxes"], CHART_COLORS):
        patch.set_facecolor(color)
        patch.set_alpha(0.8)
    ax.set_xticks(range(1, len(groups) + 1))
    ax.set_xticklabels(groups, color="#EEF2FF", fontsize=8, rotation=30, ha="right")
    ax.set_ylabel(val_col.replace("_", " "), color="#6B7DB3", fontsize=9)
    ax.tick_params(colors="#6B7DB3", labelsize=8)
    for spine in ax.spines.values(): spine.set_color("#243050")
    ax.yaxis.grid(True, color="#243050", linestyle="--", alpha=0.5)
    if title:
        ax.set_title(title, color="#EEF2FF", fontsize=11, pad=8)
    fig.tight_layout(pad=1.2)
    return fig


def render_advanced_charts_panel(df):
    """Streamlit panel with 4 advanced chart types."""
    st.markdown("""
    <div style="font-size:9px;color:#6B7DB3;font-family:monospace;
    letter-spacing:.08em;text-transform:uppercase;margin-bottom:14px">
    ◉ Advanced Visualization Studio
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs([
        "📊 Stacked Bar", "🔵 Scatter Plot", "🌡 Heatmap", "📦 Box Plot"
    ])

    with tab1:
        st.markdown("**Performance Band Distribution by Department**")
        fig = make_stacked_bar(
            df, "Department", "Performance_Band",
            title="Performance Bands by Department"
        )
        if fig:
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)

        st.markdown("**Enrollment Status by School Level**")
        fig2 = make_stacked_bar(
            df, "School_Level", "Enrollment_Status",
            title="Enrollment Status by School Level"
        )
        if fig2:
            st.pyplot(fig2, use_container_width=True)
            plt.close(fig2)

    with tab2:
        c1, c2 = st.columns(2)
        with c1:
            x_col = st.selectbox("X Axis", ["GPA", "Attendance_Pct", "Credits_Earned",
                                             "Risk_Index", "Engagement_Score"], key="sc_x")
        with c2:
            y_col = st.selectbox("Y Axis", ["Attendance_Pct", "GPA", "Credits_Earned",
                                             "Engagement_Score", "Risk_Index"], key="sc_y")
        color_col = st.selectbox("Color By", ["None", "Performance_Band", "Gender",
                                               "Department", "School_Level"], key="sc_c")
        color_col = None if color_col == "None" else color_col
        fig = make_scatter(df, x_col, y_col, color_col,
                           title=f"{x_col.replace('_',' ')} vs {y_col.replace('_',' ')}")
        if fig:
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)

    with tab3:
        r_col = st.selectbox("Row (Group)", ["Department", "Gender", "School_Level",
                                              "Performance_Band"], key="hm_r")
        c_col = st.selectbox("Column (Group)", ["Performance_Band", "Absent_Days",
                                                  "Financial_Aid_Type", "Academic_Year"], key="hm_c")
        v_col = st.selectbox("Value (Avg)", ["GPA", "Attendance_Pct",
                                               "Credits_Earned", "Risk_Index"], key="hm_v")
        fig = make_heatmap(df, r_col, c_col, v_col,
                           title=f"Avg {v_col} — {r_col} × {c_col}")
        if fig:
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)

    with tab4:
        grp = st.selectbox("Group By", ["Department", "Gender", "School_Level",
                                         "Performance_Band", "Academic_Year"], key="bp_g")
        val = st.selectbox("Metric", ["GPA", "Attendance_Pct",
                                       "Credits_Earned", "Risk_Index"], key="bp_v")
        fig = make_box_plot(df, grp, val, title=f"{val.replace('_',' ')} by {grp.replace('_',' ')}")
        if fig:
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)


# ══════════════════════════════════════════════════════════════════════════════
# NEW FEATURE 3A — EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def generate_excel_report(df, data_stats):
    """
    Generate a styled multi-sheet Excel workbook.
    Returns bytes (xlsx).
    """
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()

    # ── Color palette ──
    DARK_BG   = "0B0F1A"
    CARD_BG   = "1A2235"
    AMBER     = "F59E0B"
    GREEN     = "10B981"
    RED       = "EF4444"
    BLUE      = "3B82F6"
    PURPLE    = "8B5CF6"
    LIGHT     = "EEF2FF"
    MID       = "6B7DB3"
    BORDER_C  = "243050"

    def _amber_fill():
        return PatternFill("solid", fgColor=AMBER)
    def _dark_fill():
        return PatternFill("solid", fgColor="111827")
    def _card_fill():
        return PatternFill("solid", fgColor=CARD_BG)
    def _header_font(sz=11):
        return Font(name="Arial", bold=True, color=DARK_BG, size=sz)
    def _body_font(color=LIGHT, sz=10):
        return Font(name="Arial", color=color, size=sz)
    def _thin_border():
        s = Side(style="thin", color=BORDER_C)
        return Border(left=s, right=s, top=s, bottom=s)
    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def _left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ══════════════════════════════════════════════
    # SHEET 1 — Dashboard / KPI Summary
    # ══════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📊 Dashboard"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = AMBER

    # Title row
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "🎓  EduQuery Explorer — Academic Analytics Report"
    ws1["A1"].font      = Font(name="Arial", bold=True, color=DARK_BG, size=16)
    ws1["A1"].fill      = _amber_fill()
    ws1["A1"].alignment = _center()
    ws1.row_dimensions[1].height = 36

    # Subtitle
    ws1.merge_cells("A2:H2")
    ws1["A2"] = "Automated report generated by EduQuery Explorer · AI-powered Academic Analytics"
    ws1["A2"].font      = Font(name="Arial", color=DARK_BG, size=10)
    ws1["A2"].fill      = PatternFill("solid", fgColor="DC8C00")
    ws1["A2"].alignment = _center()
    ws1.row_dimensions[2].height = 22

    ws1.row_dimensions[3].height = 14

    # KPI Header
    ws1.merge_cells("A4:H4")
    ws1["A4"] = "KEY PERFORMANCE INDICATORS"
    ws1["A4"].font      = Font(name="Arial", bold=True, color=AMBER, size=10)
    ws1["A4"].fill      = _dark_fill()
    ws1["A4"].alignment = _center()
    ws1.row_dimensions[4].height = 20

    kpis = [
        ("Total Students",   data_stats.get("total", 0),       "records in dataset"),
        ("Departments",      data_stats.get("depts", 0),        "academic departments"),
        ("Average GPA",      data_stats.get("avg_gpa", 0),      "scale 0 – 4.0"),
        ("Avg Attendance",   f"{data_stats.get('avg_att', 0)}%","of classes attended"),
        ("At-Risk Students", data_stats.get("at_risk", 0),      "GPA below 2.0"),
        ("Above-7 Absences", data_stats.get("above7", 0),       "high-absence flag"),
    ]

    kpi_cols = ["B", "C", "D", "E", "F", "G"]
    for col_letter, (label, value, sub) in zip(kpi_cols, kpis):
        # Label row
        cell_label = f"{col_letter}5"
        ws1[cell_label] = label
        ws1[cell_label].font      = Font(name="Arial", color=MID, size=8, bold=True)
        ws1[cell_label].fill      = _card_fill()
        ws1[cell_label].alignment = _center()
        ws1.row_dimensions[5].height = 18

        # Value row
        cell_val = f"{col_letter}6"
        ws1[cell_val] = value
        color = RED if label == "At-Risk Students" else ("F97316" if label == "Above-7 Absences" else AMBER)
        ws1[cell_val].font      = Font(name="Arial", bold=True, color=color, size=18)
        ws1[cell_val].fill      = _card_fill()
        ws1[cell_val].alignment = _center()
        ws1.row_dimensions[6].height = 36

        # Sub row
        cell_sub = f"{col_letter}7"
        ws1[cell_sub] = sub
        ws1[cell_sub].font      = Font(name="Arial", color=MID, size=8)
        ws1[cell_sub].fill      = _card_fill()
        ws1[cell_sub].alignment = _center()
        ws1.row_dimensions[7].height = 16

    ws1.row_dimensions[8].height = 14

    # ── GPA by Department table ──
    ws1.merge_cells("A9:H9")
    ws1["A9"] = "AVERAGE GPA BY DEPARTMENT"
    ws1["A9"].font      = Font(name="Arial", bold=True, color=AMBER, size=10)
    ws1["A9"].fill      = _dark_fill()
    ws1["A9"].alignment = _center()
    ws1.row_dimensions[9].height = 20

    gpa_dept = (
        df.groupby("Department")["GPA"]
        .agg(["mean", "count", "std"])
        .reset_index()
        .rename(columns={"mean": "Avg GPA", "count": "Students", "std": "Std Dev"})
        .sort_values("Avg GPA", ascending=False)
    )
    headers = ["Department", "Avg GPA", "Students", "Std Dev", "Performance"]
    for ci, h in enumerate(headers, 2):
        cell = ws1.cell(row=10, column=ci)
        cell.value     = h
        cell.font      = _header_font(9)
        cell.fill      = _amber_fill()
        cell.alignment = _center()
        cell.border    = _thin_border()
    ws1.row_dimensions[10].height = 22

    for ri, (_, row) in enumerate(gpa_dept.iterrows(), 11):
        avg = round(float(row["Avg GPA"]), 2)
        perf = "Distinction" if avg >= 3.7 else "Merit" if avg >= 3.0 else "Pass" if avg >= 2.0 else "At-Risk"
        perf_color = GREEN if avg >= 3.0 else (AMBER if avg >= 2.0 else RED)
        row_data = [row["Department"], avg, int(row["Students"]),
                    round(float(row["Std Dev"]), 2), perf]
        for ci, val in enumerate(row_data, 2):
            cell = ws1.cell(row=ri, column=ci)
            cell.value     = val
            cell.font      = Font(name="Arial", color=LIGHT if ci != 6 else perf_color,
                                  size=9, bold=(ci == 6))
            cell.fill      = PatternFill("solid", fgColor="131C2E" if ri % 2 == 0 else CARD_BG)
            cell.alignment = _center()
            cell.border    = _thin_border()
        ws1.row_dimensions[ri].height = 18

    col_widths_ws1 = {"A": 4, "B": 26, "C": 12, "D": 12, "E": 12, "F": 16, "G": 4, "H": 4}
    for col, width in col_widths_ws1.items():
        ws1.column_dimensions[col].width = width

    # ══════════════════════════════════════════════
    # SHEET 2 — Full Student Data
    # ══════════════════════════════════════════════
    ws2 = wb.create_sheet("📋 Student Data")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = BLUE

    export_cols = ["Stu_ID", "Name", "Gender", "Department", "GPA",
                   "Attendance_Pct", "Credits_Earned", "Absent_Days",
                   "Performance_Band", "Risk_Index", "Engagement_Score",
                   "Academic_Year", "Financial_Aid_Type"]
    export_cols = [c for c in export_cols if c in df.columns]

    # Title
    ws2.merge_cells(f"A1:{get_column_letter(len(export_cols))}1")
    ws2["A1"] = "Complete Student Dataset"
    ws2["A1"].font      = _header_font(13)
    ws2["A1"].fill      = PatternFill("solid", fgColor=BLUE)
    ws2["A1"].alignment = _center()
    ws2.row_dimensions[1].height = 28

    # Headers
    for ci, col in enumerate(export_cols, 1):
        cell = ws2.cell(row=2, column=ci)
        cell.value     = col.replace("_", " ")
        cell.font      = _header_font(9)
        cell.fill      = _amber_fill()
        cell.alignment = _center()
        cell.border    = _thin_border()
    ws2.row_dimensions[2].height = 20

    # Data rows
    for ri, (_, row) in enumerate(df[export_cols].iterrows(), 3):
        for ci, col in enumerate(export_cols, 1):
            val  = row[col]
            cell = ws2.cell(row=ri, column=ci)
            cell.value     = val if not isinstance(val, (np.integer, np.floating)) else val.item()
            cell.alignment = _left()
            cell.border    = _thin_border()
            bg = "131C2E" if ri % 2 == 0 else CARD_BG
            cell.fill = PatternFill("solid", fgColor=bg)
            # Color code Performance_Band
            if col == "Performance_Band":
                c_map = {"Distinction": GREEN, "Merit": BLUE, "Pass": AMBER, "At-Risk": RED}
                cell.font = Font(name="Arial", color=c_map.get(str(val), LIGHT),
                                 bold=True, size=9)
            else:
                cell.font = _body_font(sz=9)

    for ci, col in enumerate(export_cols, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = max(14, len(col) + 2)
    ws2.freeze_panes = "A3"

    # ══════════════════════════════════════════════
    # SHEET 3 — At-Risk Report
    # ══════════════════════════════════════════════
    ws3 = wb.create_sheet("⚠️ At-Risk Report")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = RED

    at_risk_df = df[df["Performance_Band"] == "At-Risk"].sort_values("GPA")
    risk_cols  = ["Stu_ID", "Name", "Department", "GPA",
                  "Attendance_Pct", "Absent_Days", "Risk_Index",
                  "Engagement_Score", "Parent_Survey", "School_Satisfaction"]
    risk_cols  = [c for c in risk_cols if c in df.columns]

    ws3.merge_cells(f"A1:{get_column_letter(len(risk_cols))}1")
    ws3["A1"] = f"⚠️  At-Risk Students Report  ({len(at_risk_df)} students)"
    ws3["A1"].font      = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    ws3["A1"].fill      = PatternFill("solid", fgColor=RED)
    ws3["A1"].alignment = _center()
    ws3.row_dimensions[1].height = 32

    for ci, col in enumerate(risk_cols, 1):
        cell = ws3.cell(row=2, column=ci)
        cell.value     = col.replace("_", " ")
        cell.font      = _header_font(9)
        cell.fill      = _amber_fill()
        cell.alignment = _center()
        cell.border    = _thin_border()
    ws3.row_dimensions[2].height = 20

    for ri, (_, row) in enumerate(at_risk_df[risk_cols].iterrows(), 3):
        for ci, col in enumerate(risk_cols, 1):
            val  = row[col]
            cell = ws3.cell(row=ri, column=ci)
            cell.value     = val if not isinstance(val, (np.integer, np.floating)) else val.item()
            cell.fill      = PatternFill("solid", fgColor="1E0A0A" if ri % 2 == 0 else "2A0D0D")
            cell.alignment = _left()
            cell.border    = _thin_border()
            cell.font      = _body_font(sz=9)
    for ci, col in enumerate(risk_cols, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = max(14, len(col) + 2)
    ws3.freeze_panes = "A3"

    # ══════════════════════════════════════════════
    # SHEET 4 — Department Analytics
    # ══════════════════════════════════════════════
    ws4 = wb.create_sheet("🏫 Department Analytics")
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = GREEN

    dept_analytics = df.groupby("Department").agg(
        Total_Students=("Stu_ID", "count"),
        Avg_GPA=("GPA", "mean"),
        Avg_Attendance=("Attendance_Pct", "mean"),
        Avg_Credits=("Credits_Earned", "mean"),
        Avg_Risk_Index=("Risk_Index", "mean"),
        Avg_Engagement=("Engagement_Score", "mean"),
        At_Risk_Count=("Performance_Band", lambda x: (x == "At-Risk").sum()),
        Distinction_Count=("Performance_Band", lambda x: (x == "Distinction").sum()),
    ).reset_index()
    dept_analytics["At_Risk_Pct"] = (
        dept_analytics["At_Risk_Count"] / dept_analytics["Total_Students"] * 100
    ).round(1)
    for col in ["Avg_GPA", "Avg_Attendance", "Avg_Credits", "Avg_Risk_Index", "Avg_Engagement"]:
        dept_analytics[col] = dept_analytics[col].round(2)
    dept_analytics = dept_analytics.sort_values("Avg_GPA", ascending=False)

    ws4.merge_cells(f"A1:{get_column_letter(len(dept_analytics.columns))}1")
    ws4["A1"] = "Department-Level Analytics Summary"
    ws4["A1"].font      = Font(name="Arial", bold=True, color=DARK_BG, size=14)
    ws4["A1"].fill      = PatternFill("solid", fgColor=GREEN)
    ws4["A1"].alignment = _center()
    ws4.row_dimensions[1].height = 30

    for ci, col in enumerate(dept_analytics.columns, 1):
        cell = ws4.cell(row=2, column=ci)
        cell.value     = col.replace("_", " ")
        cell.font      = _header_font(9)
        cell.fill      = _amber_fill()
        cell.alignment = _center()
        cell.border    = _thin_border()
    ws4.row_dimensions[2].height = 22

    for ri, (_, row) in enumerate(dept_analytics.iterrows(), 3):
        for ci, col in enumerate(dept_analytics.columns, 1):
            val  = row[col]
            cell = ws4.cell(row=ri, column=ci)
            cell.value     = val if not isinstance(val, (np.integer, np.floating)) else round(val.item(), 2)
            cell.fill      = PatternFill("solid", fgColor="0E1C14" if ri % 2 == 0 else "122018")
            cell.alignment = _center()
            cell.border    = _thin_border()
            cell.font      = _body_font(sz=9)
    for ci, col in enumerate(dept_analytics.columns, 1):
        ws4.column_dimensions[get_column_letter(ci)].width = max(15, len(col) + 2)
    ws4.freeze_panes = "A3"

    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# NEW FEATURE 3B — PDF EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def _save_fig_to_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    buf.seek(0)
    return buf


def generate_pdf_report(df, data_stats):
    """Generate a styled A4 PDF report. Returns bytes."""
    buf = io.BytesIO()

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title="EduQuery Explorer — Academic Analytics Report",
    )

    styles = getSampleStyleSheet()
    W = A4[0] - 3.6*cm  # usable width

    # ── Custom styles ──
    def S(name, parent="Normal", **kw):
        return ParagraphStyle(name, parent=styles[parent], **kw)

    title_s   = S("T",  fontSize=20, textColor=colors.HexColor("#F59E0B"),
                  spaceAfter=4, alignment=TA_CENTER, fontName="Helvetica-Bold")
    sub_s     = S("Su", fontSize=10, textColor=colors.HexColor("#6B7DB3"),
                  spaceAfter=14, alignment=TA_CENTER)
    h1_s      = S("H1", fontSize=13, textColor=colors.HexColor("#EEF2FF"),
                  spaceBefore=14, spaceAfter=6, fontName="Helvetica-Bold")
    h2_s      = S("H2", fontSize=10, textColor=colors.HexColor("#F59E0B"),
                  spaceBefore=8, spaceAfter=4, fontName="Helvetica-Bold")
    body_s    = S("B",  fontSize=9,  textColor=colors.HexColor("#EEF2FF"),
                  spaceAfter=4, leading=14)
    caption_s = S("C",  fontSize=8,  textColor=colors.HexColor("#6B7DB3"),
                  alignment=TA_CENTER, spaceAfter=6)

    BG    = colors.HexColor("#0B0F1A")
    CARD  = colors.HexColor("#1A2235")
    AMBER = colors.HexColor("#F59E0B")
    GREEN = colors.HexColor("#10B981")
    RED   = colors.HexColor("#EF4444")
    BLUE  = colors.HexColor("#3B82F6")
    LIGHT = colors.HexColor("#EEF2FF")
    MID   = colors.HexColor("#6B7DB3")
    BORDER = colors.HexColor("#243050")

    story = []

    # ── Cover ──
    story.append(Spacer(1, 0.8*cm))
    story.append(Paragraph("🎓  EduQuery Explorer", title_s))
    story.append(Paragraph("Academic Analytics Report", title_s))
    story.append(Paragraph(
        "AI-powered academic dataset analysis · Intent Detection + Pattern Matching NLP",
        sub_s
    ))
    story.append(HRFlowable(width=W, thickness=1, color=AMBER, spaceAfter=14))

    # ── KPI Table ──
    story.append(Paragraph("KEY PERFORMANCE INDICATORS", h1_s))
    kpi_data = [
        ["Metric", "Value", "Description"],
        ["Total Students",   f"{data_stats.get('total', 0):,}",  "Records in dataset"],
        ["Departments",      str(data_stats.get("depts", 0)),     "Academic departments"],
        ["Average GPA",      str(data_stats.get("avg_gpa", 0)),   "Scale 0 – 4.0"],
        ["Avg Attendance",   f"{data_stats.get('avg_att', 0)}%",  "Classes attended"],
        ["At-Risk Students", str(data_stats.get("at_risk", 0)),   "GPA below 2.0"],
        ["Above-7 Absences", str(data_stats.get("above7", 0)),    "High-absence flag"],
    ]
    kpi_ts = TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), AMBER),
        ("TEXTCOLOR",   (0, 0), (-1, 0), BG),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [CARD, colors.HexColor("#131C2E")]),
        ("TEXTCOLOR",   (0, 1), (-1, -1), LIGHT),
        ("TEXTCOLOR",   (1, 5), (1, 5), RED),
        ("TEXTCOLOR",   (1, 6), (1, 6), colors.HexColor("#F97316")),
        ("GRID",        (0, 0), (-1, -1), 0.5, BORDER),
        ("ROWHEIGHT",   (0, 0), (-1, -1), 22),
    ])
    kpi_table = Table(kpi_data, colWidths=[W*0.35, W*0.2, W*0.45])
    kpi_table.setStyle(kpi_ts)
    story.append(kpi_table)
    story.append(Spacer(1, 0.4*cm))

    # ── GPA by Dept Chart ──
    story.append(Paragraph("AVERAGE GPA BY DEPARTMENT", h1_s))
    gpa_dept = (
        df.groupby("Department")["GPA"].mean()
        .reset_index().rename(columns={"GPA": "value", "Department": "label"})
        .sort_values("value", ascending=False)
    )
    n = len(gpa_dept)
    colors_bar = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(n)]
    fig, ax = plt.subplots(figsize=(7, max(3.5, n * 0.42 + 1)))
    fig.patch.set_facecolor("#111827")
    ax.set_facecolor("#1A2235")
    ax.barh(range(n), gpa_dept["value"].values[::-1],
            color=colors_bar[::-1], edgecolor="none", height=0.6)
    ax.set_yticks(range(n))
    ax.set_yticklabels(gpa_dept["label"].values[::-1], color="#EEF2FF", fontsize=8)
    ax.set_xlabel("Average GPA", color="#6B7DB3", fontsize=8)
    ax.tick_params(colors="#6B7DB3", labelsize=8)
    for spine in ax.spines.values(): spine.set_color("#243050")
    ax.xaxis.grid(True, color="#243050", linestyle="--", alpha=0.5)
    fig.tight_layout(pad=1)
    img_buf = _save_fig_to_bytes(fig)
    plt.close(fig)
    story.append(RLImage(img_buf, width=W, height=min(W * 0.55, 10*cm)))
    story.append(Paragraph("Figure 1 — Average GPA per department (sorted descending)", caption_s))

    # ── Attendance Chart ──
    story.append(Paragraph("ATTENDANCE DISTRIBUTION", h1_s))
    att_bands = [("<70%", 0, 70), ("70–80%", 70, 80), ("80–90%", 80, 90), ("90–100%", 90, 101)]
    att_df = pd.DataFrame([
        {"label": n, "value": int(df[(df["Attendance_Pct"] >= lo) & (df["Attendance_Pct"] < hi)].shape[0])}
        for n, lo, hi in att_bands
    ])
    fig, ax = plt.subplots(figsize=(6, 3))
    fig.patch.set_facecolor("#111827")
    ax.set_facecolor("#1A2235")
    ax.bar(att_df["label"], att_df["value"],
           color=["#EF4444", "#F97316", "#F59E0B", "#10B981"], edgecolor="none", width=0.6)
    ax.set_ylabel("Students", color="#6B7DB3", fontsize=8)
    ax.tick_params(colors="#EEF2FF", labelsize=8)
    for spine in ax.spines.values(): spine.set_color("#243050")
    ax.yaxis.grid(True, color="#243050", linestyle="--", alpha=0.5)
    fig.tight_layout(pad=1)
    img_buf2 = _save_fig_to_bytes(fig)
    plt.close(fig)
    story.append(RLImage(img_buf2, width=W * 0.7, height=5*cm))
    story.append(Paragraph("Figure 2 — Attendance percentage banded distribution", caption_s))

    # ── Page break ──
    story.append(PageBreak())

    # ── Performance Band Table ──
    story.append(Paragraph("PERFORMANCE BAND SUMMARY", h1_s))
    pb_summary = df["Performance_Band"].value_counts().reset_index()
    pb_summary.columns = ["Band", "Count"]
    pb_summary["% of Total"] = (pb_summary["Count"] / len(df) * 100).round(1).astype(str) + "%"
    pb_summary["Avg GPA"] = pb_summary["Band"].map(
        df.groupby("Performance_Band")["GPA"].mean().round(2)
    )
    pb_data = [["Band", "Count", "% of Total", "Avg GPA"]] + pb_summary.values.tolist()
    band_color_map = {
        "Distinction": colors.HexColor("#10B981"),
        "Merit":       colors.HexColor("#3B82F6"),
        "Pass":        colors.HexColor("#F59E0B"),
        "At-Risk":     colors.HexColor("#EF4444"),
    }
    pb_ts = TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), AMBER),
        ("TEXTCOLOR",   (0, 0), (-1, 0), BG),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [CARD, colors.HexColor("#131C2E")]),
        ("TEXTCOLOR",   (0, 1), (-1, -1), LIGHT),
        ("GRID",        (0, 0), (-1, -1), 0.5, BORDER),
        ("ROWHEIGHT",   (0, 0), (-1, -1), 22),
    ])
    pb_table = Table(pb_data, colWidths=[W*0.3, W*0.2, W*0.25, W*0.25])
    pb_table.setStyle(pb_ts)
    story.append(pb_table)
    story.append(Spacer(1, 0.4*cm))

    # ── At-Risk Dept Chart ──
    story.append(Paragraph("AT-RISK STUDENTS BY DEPARTMENT", h1_s))
    risk_dept = (
        df[df["Performance_Band"] == "At-Risk"]
        .groupby("Department").size()
        .reset_index().rename(columns={0: "value", "Department": "label"})
        .sort_values("value", ascending=False)
    )
    if not risk_dept.empty:
        fig, ax = plt.subplots(figsize=(7, max(3, len(risk_dept) * 0.4 + 1)))
        fig.patch.set_facecolor("#111827")
        ax.set_facecolor("#1A2235")
        ax.barh(range(len(risk_dept)), risk_dept["value"].values[::-1],
                color="#EF4444", edgecolor="none", height=0.6)
        ax.set_yticks(range(len(risk_dept)))
        ax.set_yticklabels(risk_dept["label"].values[::-1], color="#EEF2FF", fontsize=8)
        ax.set_xlabel("At-Risk Student Count", color="#6B7DB3", fontsize=8)
        ax.tick_params(colors="#6B7DB3", labelsize=8)
        for spine in ax.spines.values(): spine.set_color("#243050")
        ax.xaxis.grid(True, color="#243050", linestyle="--", alpha=0.5)
        fig.tight_layout(pad=1)
        img_buf3 = _save_fig_to_bytes(fig)
        plt.close(fig)
        story.append(RLImage(img_buf3, width=W, height=min(W * 0.5, 9*cm)))
        story.append(Paragraph("Figure 3 — At-Risk student count per department", caption_s))

    # ── Correlation Table ──
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("CORRELATION ANALYSIS", h1_s))
    pairs = [
        ("GPA", "Attendance_Pct"),
        ("GPA", "Credits_Earned"),
        ("Attendance_Pct", "Credits_Earned"),
        ("Risk_Index", "GPA"),
        ("Engagement_Score", "GPA"),
    ]
    def _pearson(a, b):
        x = pd.to_numeric(df[a], errors="coerce")
        y = pd.to_numeric(df[b], errors="coerce")
        mask = x.notna() & y.notna()
        if mask.sum() < 2: return "N/A"
        r, _ = stats.pearsonr(x[mask], y[mask])
        return round(r, 3)
    def _strength(r):
        if r == "N/A": return "—"
        ar = abs(float(r))
        return "Strong" if ar > 0.5 else "Moderate" if ar > 0.3 else "Weak"

    corr_data = [["Variable A", "Variable B", "Pearson r", "Strength"]]
    for a, b in pairs:
        r = _pearson(a, b)
        corr_data.append([a.replace("_"," "), b.replace("_"," "), str(r), _strength(r)])
    corr_ts = TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0), AMBER),
        ("TEXTCOLOR",      (0, 0), (-1, 0), BG),
        ("FONTNAME",       (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, -1), 9),
        ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [CARD, colors.HexColor("#131C2E")]),
        ("TEXTCOLOR",      (0, 1), (-1, -1), LIGHT),
        ("GRID",           (0, 0), (-1, -1), 0.5, BORDER),
        ("ROWHEIGHT",      (0, 0), (-1, -1), 22),
    ])
    corr_table = Table(corr_data, colWidths=[W*0.28, W*0.28, W*0.2, W*0.24])
    corr_table.setStyle(corr_ts)
    story.append(corr_table)

    # ── Footer ──
    story.append(Spacer(1, 0.6*cm))
    story.append(HRFlowable(width=W, thickness=0.5, color=BORDER, spaceAfter=6))
    story.append(Paragraph(
        "Generated by EduQuery Explorer · AI-powered Academic Analytics Platform",
        S("foot", fontSize=8, textColor=MID, alignment=TA_CENTER)
    ))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# ORIGINAL NLP & QUERY ENGINE (unchanged)
# ══════════════════════════════════════════════════════════════════════════════
INTENT_CONFIG = [
    {"id": "correlation",  "weight": 20,
     "patterns": [r"\b(correlat|correlation|relation|impact|affect|influence|link)\b"]},
    {"id": "risk",         "weight": 20,
     "patterns": [r"\bat.risk\b", r"\brisk\b", r"\bstruggling\b"]},
    {"id": "count",        "weight": 20,
     "patterns": [r"\b(how many|count|number of)\b"]},
    {"id": "distribution", "weight": 20,
     "patterns": [r"\b(distribution|spread|breakdown|split|proportion|percentage)\b"]},
    {"id": "rank",         "weight": 20,
     "patterns": [r"\btop\s+\d+\b", r"\bbottom\s+\d+\b",
                  r"\b(rank|highest|lowest)\b"]},
    {"id": "average",      "weight": 15,
     "patterns": [r"\b(average|mean|avg)\b"]},
    {"id": "compare",      "weight": 15,
     "patterns": [r"\b(compare|vs\.?|versus|difference|between)\b"]},
    {"id": "filter",       "weight": 10,
     "patterns": [r"\b(filter|find|list|who)\b",
                  r"students\s+with\b", r"\bwhere\b"]},
    {"id": "summary",      "weight":  5,
     "patterns": [r"\b(summar|overview|insight|report|describe|about|tell me)\b"]},
]


def detect_intent(prompt):
    p = prompt.lower()
    scores = {}
    matched = []
    for cfg in INTENT_CONFIG:
        w = cfg["weight"]
        for pat in cfg["patterns"]:
            if re.search(pat, p, re.IGNORECASE):
                scores[cfg["id"]] = scores.get(cfg["id"], 0) + w
                matched.append({"intent": cfg["id"], "pattern": pat})
    if re.search(r"\bshow\b.*\bby\b", p) and "filter" in scores and "average" not in scores:
        scores["average"] = scores.get("average", 0) + 15
    best = sorted(scores.items(), key=lambda x: -x[1])
    intent = best[0][0] if best else "summary"
    return {"intent": intent, "scores": scores, "matched": matched[:4]}


FIELD_MAP = {
    "GPA":              [r"\bgpa\b", r"\bgrade.?point\b", r"\bgrades?\b"],
    "Attendance_Pct":   [r"\battendance\b", r"\battend\b", r"\bpresent\b"],
    "Credits_Earned":   [r"\bcredits?\b", r"\bcredits.earned\b"],
    "Absent_Days":      [r"\babsent\b", r"\babsence\b"],
    "Risk_Index":       [r"\brisk\b", r"\bat.risk\b"],
    "Engagement_Score": [r"\bengagement\b"],
}

GROUPBY_MAP = {
    "Department":          [r"by\s+department", r"per\s+department",
                            r"each\s+department", r"\bdepartment\b"],
    "Gender":              [r"by\s+gender", r"per\s+gender",
                            r"\bgender\b", r"\bsex\b"],
    "School_Level":        [r"school\s+level", r"by\s+school",
                            r"lower school|middle school|high school"],
    "Enrollment_Status":   [r"enrollment", r"full.?time|part.?time"],
    "Financial_Aid_Type":  [r"financial.?aid", r"aid\s+type",
                            r"scholarship", r"\baid\b"],
    "School_Satisfaction": [r"satisfaction", r"satisfied"],
    "Performance_Band":    [r"performance.?band", r"\bband\b"],
    "Academic_Year":       [r"academic.?year",
                            r"freshman|sophomore|junior|senior"],
    "Absent_Days":         [r"absent.?days", r"absence\s+categor"],
    "State":               [r"by\s+state", r"\bstate\b"],
    "Parent_Survey":       [r"parent.?survey"],
}


def pattern_match(prompt):
    p = prompt.lower()
    fields = [
        field for field, patterns in FIELD_MAP.items()
        if any(re.search(pat, p, re.IGNORECASE) for pat in patterns)
    ]
    group_by = None
    for col, patterns in GROUPBY_MAP.items():
        if any(re.search(pat, p, re.IGNORECASE) for pat in patterns):
            group_by = col
            break
    filters = []
    if re.search(r"high\s+gpa|top\s+gpa", p):
        filters.append({"key": "GPA", "op": "gte", "val": 3.5})
    m = re.search(r"gpa\s*[>≥>=]{1,2}\s*([\d.]+)", p)
    if m:
        filters.append({"key": "GPA", "op": "gte", "val": float(m.group(1))})
    m = re.search(r"gpa\s*[<≤<=]{1,2}\s*([\d.]+)", p)
    if m:
        filters.append({"key": "GPA", "op": "lte", "val": float(m.group(1))})
    if re.search(r"low\s+attendance", p):
        filters.append({"key": "Attendance_Pct", "op": "lte", "val": 75.0})
    m = re.search(r"attendance\s*[>≥>=]{1,2}\s*([\d.]+)", p)
    if m:
        filters.append({"key": "Attendance_Pct", "op": "gte", "val": float(m.group(1))})
    if re.search(r"above.?7|absent.+above|more\s+than\s+7\s+absent", p):
        filters.append({"key": "Absent_Days", "op": "eq", "val": "Above-7"})
    if re.search(r"under.?7|below.?7", p):
        filters.append({"key": "Absent_Days", "op": "eq", "val": "Under-7"})
    if re.search(r"\bfull.?time\b", p):
        filters.append({"key": "Enrollment_Status", "op": "eq", "val": "Full-time"})
    if re.search(r"\bpart.?time\b", p):
        filters.append({"key": "Enrollment_Status", "op": "eq", "val": "Part-time"})
    if re.search(r"\bfemale\b", p):
        filters.append({"key": "Gender", "op": "eq", "val": "Female"})
    if re.search(r"\bmale\b", p) and not re.search(r"\bfemale\b", p):
        filters.append({"key": "Gender", "op": "eq", "val": "Male"})
    if re.search(r"non.?binary", p):
        filters.append({"key": "Gender", "op": "eq", "val": "Non-binary"})
    if re.search(r"at.risk", p):
        filters.append({"key": "Performance_Band", "op": "eq", "val": "At-Risk"})
    if re.search(r"\bdistinction\b", p):
        filters.append({"key": "Performance_Band", "op": "eq", "val": "Distinction"})
    for dept in [
        "Engineering", "Physics", "Mathematics", "Chemistry", "History",
        "Economics", "Sociology", "Psychology", "Political Science",
        "Business Administration", "Computer Science", "Biology",
        "English Literature", "Art & Design", "Education",
    ]:
        if re.search(re.escape(dept), p, re.IGNORECASE):
            filters.append({"key": "Department", "op": "eq", "val": dept})
    for lvl in ["Lower School", "Middle School", "High School"]:
        if re.search(re.escape(lvl), p, re.IGNORECASE):
            filters.append({"key": "School_Level", "op": "eq", "val": lvl})
    nm = re.search(r"\btop\s+(\d+)\b|\bbottom\s+(\d+)\b|\bfirst\s+(\d+)\b", p, re.IGNORECASE)
    top_n = int(nm.group(1) or nm.group(2) or nm.group(3)) if nm else 5
    is_bottom = bool(re.search(r"\bbottom\b|\bworst\b|\blowest\b", p, re.IGNORECASE))
    return {"fields": fields, "group_by": group_by, "filters": filters,
            "top_n": top_n, "is_bottom": is_bottom}


def apply_filters(df, filters):
    result = df.copy()
    for f in filters:
        key, op, val = f["key"], f["op"], f["val"]
        if key not in result.columns:
            continue
        if op == "eq":
            result = result[result[key].astype(str) == str(val)]
        elif op == "gte":
            result = result[pd.to_numeric(result[key], errors="coerce") >= float(val)]
        elif op == "lte":
            result = result[pd.to_numeric(result[key], errors="coerce") <= float(val)]
    return result


def group_avg(df, group_col, metric_col):
    if group_col not in df.columns or metric_col not in df.columns:
        return pd.DataFrame()
    grp = (
        df.groupby(group_col)[metric_col]
        .agg(["mean", "count"])
        .reset_index()
        .rename(columns={"mean": "value", group_col: "label"})
    )
    grp["value"] = grp["value"].round(2)
    return grp.sort_values("value", ascending=False)


def count_by(df, col):
    if col not in df.columns:
        return pd.DataFrame()
    counts = df[col].value_counts().reset_index()
    counts.columns = ["label", "value"]
    total = len(df)
    counts["pct"] = (counts["value"] / total * 100).round(1)
    return counts


def num_stats(df, col):
    if col not in df.columns:
        return {}
    vals = pd.to_numeric(df[col], errors="coerce").dropna()
    if vals.empty:
        return {}
    return {
        "avg":    round(vals.mean(), 2),
        "median": round(vals.median(), 2),
        "min":    round(vals.min(), 2),
        "max":    round(vals.max(), 2),
        "std":    round(vals.std(), 2),
        "count":  len(vals),
    }


def pearson_r(df, col_a, col_b):
    a = pd.to_numeric(df[col_a], errors="coerce")
    b = pd.to_numeric(df[col_b], errors="coerce")
    mask = a.notna() & b.notna()
    if mask.sum() < 2:
        return 0.0
    r, _ = stats.pearsonr(a[mask], b[mask])
    return round(float(r), 3)


def execute_query(df, prompt):
    nlp1 = detect_intent(prompt)
    nlp2 = pattern_match(prompt)
    intent    = nlp1["intent"]
    fields    = nlp2["fields"]
    group_by  = nlp2["group_by"]
    filters   = nlp2["filters"]
    top_n     = nlp2["top_n"]
    is_bottom = nlp2["is_bottom"]

    primary_metric = (
        fields[0] if fields
        else ("Risk_Index" if intent == "risk" else "GPA")
    )
    eff_group = group_by or (
        "Gender" if intent == "compare" and "Gender" in fields else "Department"
    )

    chart_type   = "bar"
    chart_df     = pd.DataFrame()
    title        = ""
    insight      = ""
    sample_table = pd.DataFrame()

    if intent == "average":
        if group_by:
            chart_df = group_avg(df, group_by, primary_metric)
            title    = f"Average {primary_metric.replace('_',' ')} by {group_by.replace('_',' ')}"
            if not chart_df.empty:
                best = chart_df.iloc[0]; worst = chart_df.iloc[-1]
                insight = (f"Highest: **{best['label']}** ({best['value']}) | "
                           f"Lowest: **{worst['label']}** ({worst['value']}) | "
                           f"Spread: {round(float(best['value'])-float(worst['value']),2)}")
        else:
            gs = num_stats(df, "GPA"); as_ = num_stats(df, "Attendance_Pct"); cs = num_stats(df, "Credits_Earned")
            chart_df = pd.DataFrame({"label": ["Avg GPA","Avg Attendance %","Avg Credits"],
                                     "value": [gs.get("avg",0), as_.get("avg",0), cs.get("avg",0)]})
            title   = "Overall Dataset Averages"
            insight = (f"Avg GPA: **{gs.get('avg',0)}** | Avg Attendance: **{as_.get('avg',0)}%** | "
                       f"Avg Credits: **{cs.get('avg',0)}**")
    elif intent == "compare":
        chart_df = group_avg(df, eff_group, primary_metric)
        title    = f"Comparing {primary_metric.replace('_',' ')} by {eff_group.replace('_',' ')}"
        if not chart_df.empty:
            best = chart_df.iloc[0]; worst = chart_df.iloc[-1]
            insight = (f"**{best['label']}** leads ({best['value']}); "
                       f"**{worst['label']}** lowest ({worst['value']}). "
                       f"Gap: {round(float(best['value'])-float(worst['value']),2)}")
    elif intent == "rank":
        all_data = group_avg(df, eff_group, primary_metric)
        chart_df = (all_data.sort_values("value").head(top_n) if is_bottom else all_data.head(top_n))
        lbl   = "Bottom" if is_bottom else "Top"
        title = f"{lbl} {top_n} {eff_group.replace('_',' ')} by {primary_metric.replace('_',' ')}"
        if not chart_df.empty:
            insight = (f"{'Lowest' if is_bottom else 'Highest'}: "
                       f"**{chart_df.iloc[0]['label']}** ({chart_df.iloc[0]['value']}) | "
                       f"Range: {chart_df['value'].min()} – {chart_df['value'].max()}")
    elif intent == "count":
        col = group_by or ("Gender" if "Gender" in fields else
                           "Absent_Days" if "Absent_Days" in fields else "Department")
        chart_df   = count_by(df, col)
        chart_type = "pie" if len(chart_df) <= 7 else "bar"
        title      = f"Student Count by {col.replace('_',' ')} (Total: {len(df):,})"
        if not chart_df.empty:
            top = chart_df.iloc[0]
            insight = f"**{top['label']}** is the largest group: {top['value']} students ({top['pct']}%)"
    elif intent == "distribution":
        if "GPA" in fields or re.search(r"\bgpa\b", prompt, re.IGNORECASE):
            bands = [("0.0–1.0",0,1),("1.0–2.0",1,2),("2.0–3.0",2,3),("3.0–4.0",3,4.1)]
            chart_df = pd.DataFrame([{"label": n,
                                       "value": int(df[(df["GPA"]>=lo)&(df["GPA"]<hi)].shape[0])}
                                      for n,lo,hi in bands])
            title = "GPA Distribution (Banded)"
            risk_n = int((df["GPA"] < 2.0).sum())
            insight = f"**{risk_n}** students ({round(risk_n/len(df)*100,1)}%) in At-Risk band (GPA < 2.0)"
        elif re.search(r"attendance", prompt, re.IGNORECASE):
            bands = [("<70%",0,70),("70–80%",70,80),("80–90%",80,90),("90–100%",90,101)]
            chart_df = pd.DataFrame([{"label": n,
                                       "value": int(df[(df["Attendance_Pct"]>=lo)&(df["Attendance_Pct"]<hi)].shape[0])}
                                      for n,lo,hi in bands])
            title = "Attendance Distribution (Banded)"
            low_n = int((df["Attendance_Pct"] < 75).sum())
            insight = f"**{low_n}** students ({round(low_n/len(df)*100,1)}%) have critically low attendance below 75%"
        else:
            col = group_by or "Performance_Band"
            chart_df   = count_by(df, col)
            chart_type = "pie"
            title      = f"Distribution by {col.replace('_',' ')}"
            if not chart_df.empty:
                insight = f"**{chart_df.iloc[0]['label']}** dominates at {chart_df.iloc[0]['pct']}% of students"
    elif intent == "correlation":
        pairs = [("GPA","Attendance_Pct"),("GPA","Credits_Earned"),
                 ("Attendance_Pct","Credits_Earned"),("Risk_Index","GPA"),
                 ("Engagement_Score","GPA")]
        chart_df = pd.DataFrame([{
            "label": (f"{a.replace('_Pct',' %').replace('_',' ')} ↔ "
                      f"{b.replace('_Pct',' %').replace('_',' ')}"),
            "value": pearson_r(df, a, b),
        } for a, b in pairs])
        title = "Pearson Correlation Between Key Metrics"
        r0 = chart_df.iloc[0]["value"] if not chart_df.empty else 0
        strength = "strong" if abs(r0) > 0.5 else ("moderate" if abs(r0) > 0.3 else "weak")
        insight = (f"GPA ↔ Attendance r = **{r0}**: {strength} "
                   f"{'positive' if r0 >= 0 else 'negative'} relationship. "
                   f"Higher attendance predicts better academic outcomes.")
    elif intent == "risk":
        risk_df  = df[(df["Performance_Band"] == "At-Risk") | (df["GPA"] < 2.0)]
        chart_df = count_by(risk_df, "Department")
        title    = f"At-Risk Students by Department ({len(risk_df):,} total)"
        if not chart_df.empty:
            insight = (f"**{len(risk_df)}** students "
                       f"({round(len(risk_df)/len(df)*100,1)}%) are academically at-risk. "
                       f"Most in **{chart_df.iloc[0]['label']}** ({chart_df.iloc[0]['value']})")
    elif intent == "filter":
        filtered = apply_filters(df, filters) if filters else df.head(20)
        chart_df = count_by(filtered, "Department")
        title    = f"Filter Results: {len(filtered):,} students matched"
        sample_table = filtered[[
            "Stu_ID","Name","Gender","Department","GPA",
            "Attendance_Pct","Absent_Days","Performance_Band",
        ]].head(10).copy()
        if not chart_df.empty:
            insight = (f"**{len(filtered)}** students matched the filter criteria. "
                       f"Top dept in results: **{chart_df.iloc[0]['label']}** ({chart_df.iloc[0]['value']})")
        else:
            insight = f"**{len(filtered)}** students matched."
    else:
        chart_df   = count_by(df, "Department").head(10)
        chart_type = "bar"
        title      = f"Dataset Overview — {len(df):,} Students by Department"
        gs = num_stats(df, "GPA"); as_ = num_stats(df, "Attendance_Pct")
        above7 = int(df["Absent_Days"].eq("Above-7").sum())
        risk_n = int(df["Performance_Band"].eq("At-Risk").sum())
        insight = (f"**{len(df):,}** students | Avg GPA: **{gs.get('avg',0)}** | "
                   f"Avg Attendance: **{as_.get('avg',0)}%** | "
                   f"Above-7 absences: **{above7}** | "
                   f"At-Risk: **{risk_n}** ({round(risk_n/len(df)*100,1)}%)")

    return {"intent": intent, "nlp1": nlp1, "nlp2": nlp2,
            "chart_type": chart_type, "chart_df": chart_df,
            "title": title, "insight": insight, "sample_table": sample_table}


# ══════════════════════════════════════════════════════════════════════════════
# ORIGINAL CHART RENDERERS (unchanged)
# ══════════════════════════════════════════════════════════════════════════════
def make_bar_chart(chart_df, title):
    if chart_df.empty: return None
    labels = chart_df["label"].astype(str).tolist()
    values = chart_df["value"].astype(float).tolist()
    n      = len(labels)
    colors = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(n)]
    fig, ax = plt.subplots(figsize=(8, max(3, n * 0.42 + 1)))
    fig.patch.set_facecolor("#111827")
    ax.set_facecolor("#1A2235")
    y_pos = range(n)
    bars = ax.barh(y_pos, values, color=colors, edgecolor="none", height=0.62, zorder=3)
    x_max = max(values) if values else 1
    for bar, val in zip(bars, values):
        ax.text(val + x_max * 0.01, bar.get_y() + bar.get_height() / 2,
                f"{val:.2f}" if isinstance(val, float) and val % 1 != 0 else f"{int(val)}",
                va="center", ha="left", color="#EEF2FF", fontsize=9, fontfamily="monospace")
    ax.set_yticks(list(y_pos))
    ax.set_yticklabels(labels, color="#EEF2FF", fontsize=9)
    ax.set_xlabel("Value", color="#6B7DB3", fontsize=9)
    ax.tick_params(colors="#6B7DB3", labelsize=9)
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#243050"); ax.spines["bottom"].set_color("#243050")
    ax.xaxis.grid(True, color="#243050", linestyle="--", alpha=0.6, zorder=0)
    ax.set_axisbelow(True)
    if any(v < 0 for v in values):
        ax.axvline(0, color="#6B7DB3", linewidth=0.8)
    ax.invert_yaxis()
    fig.tight_layout(pad=1.2)
    return fig


def make_pie_chart(chart_df, title):
    if chart_df.empty: return None
    labels = chart_df["label"].astype(str).tolist()
    values = chart_df["value"].astype(float).tolist()
    colors = [CHART_COLORS[i % len(CHART_COLORS)] for i in range(len(labels))]
    fig, ax = plt.subplots(figsize=(7, 5))
    fig.patch.set_facecolor("#111827"); ax.set_facecolor("#111827")
    wedges, texts, autotexts = ax.pie(
        values, labels=None, colors=colors, autopct="%1.1f%%",
        startangle=140, pctdistance=0.78,
        wedgeprops={"edgecolor": "#0B0F1A", "linewidth": 2})
    for at in autotexts:
        at.set_color("#0B0F1A"); at.set_fontsize(8); at.set_fontweight("bold")
    ax.add_patch(plt.Circle((0, 0), 0.55, fc="#111827"))
    ax.legend(wedges, [f"{l} ({v})" for l, v in zip(labels, values)],
              loc="center left", bbox_to_anchor=(1, 0, 0.5, 1),
              fontsize=9, frameon=False, labelcolor="#EEF2FF")
    fig.tight_layout(pad=1.2)
    return fig


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ══════════════════════════════════════════════════════════════════════════════
for key, default in [
    ("messages", []), ("df", None), ("prep_log", []),
    ("pipe_step", -1), ("data_stats", {}),
]:
    if key not in st.session_state:
        st.session_state[key] = default


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE BANNER
# ══════════════════════════════════════════════════════════════════════════════
def render_pipeline(active_step):
    ps  = st.session_state.pipe_step
    html_parts = ['<div class="pipeline-wrap">']
    for i, (icon, label) in enumerate(PIPELINE_STEPS):
        cls = "done" if ps > i else ("active" if ps == i else "pending")
        icon_show = "✓" if ps > i else icon
        html_parts.append(
            f'<div class="pipe-step {cls}"><span>{icon_show}</span> {label}</div>'
        )
        if i < len(PIPELINE_STEPS) - 1:
            html_parts.append('<span class="pipe-arrow">→</span>')
    html_parts.append("</div>")
    st.markdown("".join(html_parts), unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div style="padding:14px 0 8px 0">
      <div style="font-size:9px;color:#6B7DB3;font-family:'JetBrains Mono',monospace;
                  letter-spacing:.1em;text-transform:uppercase;margin-bottom:10px">
        01 · Dataset Upload
      </div>
    </div>""", unsafe_allow_html=True)

    uploaded_file = st.file_uploader(
        "Upload Education Dataset",
        type=["xlsx", "xls", "csv"],
        label_visibility="collapsed",
    )

    if uploaded_file is not None:
        file_key = f"{uploaded_file.name}_{uploaded_file.size}"
        if st.session_state.get("_file_key") != file_key:
            st.session_state["_file_key"] = file_key
            with st.spinner("Running preprocessing pipeline…"):
                st.session_state.pipe_step = 0
                file_bytes = uploaded_file.read()
                df_clean, prep_log = preprocess_dataframe(file_bytes, uploaded_file.name)
                st.session_state.df       = df_clean
                st.session_state.prep_log = prep_log
                st.session_state.pipe_step = 2
                risk_n = int(df_clean["Performance_Band"].eq("At-Risk").sum())
                above7 = int(df_clean["Absent_Days"].eq("Above-7").sum())
                st.session_state.data_stats = {
                    "total":   len(df_clean),
                    "columns": len(df_clean.columns),
                    "depts":   int(df_clean["Department"].nunique()),
                    "avg_gpa": round(float(df_clean["GPA"].mean()), 2),
                    "avg_att": round(float(df_clean["Attendance_Pct"].mean()), 1),
                    "at_risk": risk_n,
                    "above7":  above7,
                }
                st.session_state.messages.append({
                    "role": "system",
                    "text": (
                        f"✅ Dataset **{uploaded_file.name}** loaded & preprocessed!\n\n"
                        f"📊 **{len(df_clean):,} students** · **{df_clean['Department'].nunique()} departments** · "
                        f"**{len(df_clean.columns)} columns**\n\n"
                        f"📈 Avg GPA: **{round(float(df_clean['GPA'].mean()),2)}** | "
                        f"Avg Attendance: **{round(float(df_clean['Attendance_Pct'].mean()),1)}%**\n\n"
                        f"⚠️ At-Risk students: **{risk_n}** | Above-7 absences: **{above7}**\n\n"
                        f"*Pipeline ready — ask any question below.*"
                    ),
                })

    if st.session_state.prep_log:
        st.markdown("""<div style="font-size:9px;color:#6B7DB3;font-family:'JetBrains Mono',monospace;
                    letter-spacing:.1em;text-transform:uppercase;margin:14px 0 8px 0">
          02 · Preprocessing Log</div>""", unsafe_allow_html=True)
        log_html = ""
        for item in st.session_state.prep_log:
            log_html += f"""
            <div class="prep-item">
              <div class="prep-check">✓</div>
              <div><div class="prep-name">{item['step']}</div>
              <div class="prep-detail">{item['detail']}</div></div>
            </div>"""
        st.markdown(log_html, unsafe_allow_html=True)

    st.markdown("""<div style="font-size:9px;color:#6B7DB3;font-family:'JetBrains Mono',monospace;
                letter-spacing:.1em;text-transform:uppercase;margin:14px 0 8px 0">
      03 · NLP Modules</div>
    <div style="margin-bottom:8px;padding:10px;border-radius:8px;
          border:1px solid rgba(245,158,11,.25);background:rgba(245,158,11,.05)">
      <div style="font-size:9px;color:#F59E0B;font-family:'JetBrains Mono',monospace">MODULE 1</div>
      <div style="font-size:11px;font-weight:600;color:#EEF2FF;margin:3px 0">Intent Detection</div>
      <div style="font-size:9px;color:#6B7DB3;line-height:1.5">9 semantic intents · weighted regex scoring</div>
    </div>
    <div style="padding:10px;border-radius:8px;
          border:1px solid rgba(59,130,246,.25);background:rgba(59,130,246,.05)">
      <div style="font-size:9px;color:#3B82F6;font-family:'JetBrains Mono',monospace">MODULE 2</div>
      <div style="font-size:11px;font-weight:600;color:#EEF2FF;margin:3px 0">Pattern Matching</div>
      <div style="font-size:9px;color:#6B7DB3;line-height:1.5">Fields · groupBy · filters · topN</div>
    </div>""", unsafe_allow_html=True)

    ds = st.session_state.data_stats
    if ds:
        st.markdown("""<div style="font-size:9px;color:#6B7DB3;font-family:'JetBrains Mono',monospace;
                    letter-spacing:.1em;text-transform:uppercase;margin:14px 0 8px 0">
          04 · Dataset Stats</div>""", unsafe_allow_html=True)
        for k, v in [("Total Students", f"{ds['total']:,}"), ("Departments", ds["depts"]),
                     ("Avg GPA", ds["avg_gpa"]), ("Avg Attendance", f"{ds['avg_att']}%"),
                     ("At-Risk Students", ds["at_risk"]), ("Above-7 Absences", ds["above7"])]:
            st.markdown(
                f'<div style="display:flex;justify-content:space-between;padding:5px 0;'
                f'border-bottom:1px solid #1A2235;font-size:11px">'
                f'<span style="color:#6B7DB3">{k}</span>'
                f'<span style="color:#F59E0B;font-family:monospace;font-weight:600">{v}</span></div>',
                unsafe_allow_html=True)

    if st.session_state.df is not None:
        st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)
        if st.button("↺ Reset / New Dataset", use_container_width=True):
            for key in ["df","prep_log","messages","data_stats","pipe_step","_file_key"]:
                if key in st.session_state: del st.session_state[key]
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# MAIN AREA
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="header-banner">
  <div style="font-size:30px;background:linear-gradient(135deg,#F59E0B,#DC8C00);
              border-radius:10px;width:48px;height:48px;display:flex;
              align-items:center;justify-content:center;flex-shrink:0">🎓</div>
  <div>
    <div class="header-title">EduQuery <span style="color:#F59E0B">Explorer</span></div>
    <div class="header-sub">Prompt-Based Analysis · NLP · ML Risk Model · Advanced Charts · PDF/Excel Export</div>
  </div>
  <div style="margin-left:auto;display:flex;gap:8px;align-items:center;flex-wrap:wrap">
    <span class="badge-green">NLP · 9 intents</span>
    <span class="badge-purple">ML Risk Model</span>
    <span class="badge-amber">PDF + Excel Export</span>
  </div>
</div>
""", unsafe_allow_html=True)

render_pipeline(st.session_state.pipe_step)

if st.session_state.df is None:
    st.markdown("""
    <div style="text-align:center;padding:60px 20px">
      <div style="font-size:52px;margin-bottom:16px">📊</div>
      <div style="font-size:24px;font-weight:700;color:#EEF2FF;margin-bottom:10px">
        Upload your dataset to begin
      </div>
      <div style="color:#6B7DB3;font-size:14px;max-width:500px;margin:0 auto;line-height:1.7">
        Upload <code style="color:#FCD34D;background:rgba(245,158,11,.1);border-radius:4px;padding:1px 6px">Combined_Education_Dataset.xlsx</code>
        from the sidebar, then use the tabs below to explore NLP queries,
        the ML risk predictor, advanced charts, and export reports.
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Data loaded — show KPI cards ──
ds = st.session_state.data_stats
st.markdown(f"""
<div class="metric-row">
  <div class="metric-card"><div class="metric-label">Total Students</div>
    <div class="metric-value">{ds.get('total',0):,}</div><div class="metric-sub">records loaded</div></div>
  <div class="metric-card"><div class="metric-label">Avg GPA</div>
    <div class="metric-value">{ds.get('avg_gpa',0)}</div><div class="metric-sub">scale 0 – 4.0</div></div>
  <div class="metric-card"><div class="metric-label">Avg Attendance</div>
    <div class="metric-value">{ds.get('avg_att',0)}%</div><div class="metric-sub">of classes attended</div></div>
  <div class="metric-card"><div class="metric-label">Departments</div>
    <div class="metric-value">{ds.get('depts',0)}</div><div class="metric-sub">academic departments</div></div>
  <div class="metric-card"><div class="metric-label">At-Risk</div>
    <div class="metric-value" style="color:#EF4444">{ds.get('at_risk',0)}</div>
    <div class="metric-sub">GPA below 2.0</div></div>
  <div class="metric-card"><div class="metric-label">Above-7 Absences</div>
    <div class="metric-value" style="color:#F97316">{ds.get('above7',0)}</div>
    <div class="metric-sub">high absence flag</div></div>
</div>
""", unsafe_allow_html=True)

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN TABS — NEW LAYOUT
# ══════════════════════════════════════════════════════════════════════════════
tab_query, tab_ml, tab_charts, tab_export = st.tabs([
    "💬 NLP Query",
    "🤖 ML Risk Model",
    "📊 Advanced Charts",
    "📥 Export Reports",
])

# ─────────────────────────────────────────────
# TAB 1 — Original NLP Query (fully preserved)
# ─────────────────────────────────────────────
with tab_query:
    # Quick prompts
    st.markdown(
        '<div style="font-size:10px;color:#6B7DB3;font-family:monospace;'
        'letter-spacing:.08em;text-transform:uppercase;margin-bottom:8px">Quick Prompts</div>',
        unsafe_allow_html=True,
    )
    chip_cols = st.columns(6)
    for i, qp in enumerate(QUICK_PROMPTS[:6]):
        with chip_cols[i % 6]:
            if st.button(qp, key=f"chip_{i}"):
                st.session_state["_pending_prompt"] = qp
    chip_cols2 = st.columns(6)
    for i, qp in enumerate(QUICK_PROMPTS[6:]):
        with chip_cols2[i % 6]:
            if st.button(qp, key=f"chip2_{i}"):
                st.session_state["_pending_prompt"] = qp

    st.markdown("---")

    # Chat history
    for msg in st.session_state.messages:
        if msg["role"] == "user":
            st.markdown(f'<div class="msg-user">{msg["text"]}</div>', unsafe_allow_html=True)
        elif msg["role"] == "system":
            parts = msg["text"].split("**")
            formatted = ""
            for j, part in enumerate(parts):
                if j % 2 == 1:
                    formatted += f"<strong style='color:#FCD34D'>{part}</strong>"
                else:
                    formatted += part.replace("\n", "<br/>")
            st.markdown(
                f'<div class="msg-assistant"><div class="msg-meta">EDUQUERY</div>'
                f'<div class="msg-text">{formatted}</div></div>',
                unsafe_allow_html=True)
        elif msg["role"] == "assistant":
            r = msg["result"]
            st.markdown(
                f'<div class="msg-assistant"><div class="msg-meta">EDUQUERY · Result</div>'
                f'<div class="msg-title">{r["title"]}</div>',
                unsafe_allow_html=True)

            # NLP Trace
            nlp1 = r["nlp1"]; nlp2 = r["nlp2"]
            intent_tags = "".join([
                f'<span class="nlp-tag-intent">{m["intent"]}: /{m["pattern"]}/</span>'
                for m in nlp1.get("matched", [])[:3]])
            field_tags = "".join([f'<span class="nlp-tag-field">{f}</span>'
                                   for f in nlp2.get("fields", [])])
            group_tag = (f'<span class="nlp-tag-group">groupBy: {nlp2["group_by"]}</span>'
                         if nlp2.get("group_by") else "")
            filter_tags = "".join([
                f'<span class="nlp-tag-filter">{f["key"]} {f["op"]} {f["val"]}</span>'
                for f in nlp2.get("filters", [])])
            st.markdown(f"""
            <div class="nlp-trace">
              <div class="nlp-title">◈ NLP Analysis Trace</div>
              <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px">
                <div><div style="font-size:9px;color:#6B7DB3;margin-bottom:4px">Module 1 — Intent Detection</div>
                  <span class="nlp-tag-intent" style="font-weight:700;font-size:10px">INTENT: {nlp1.get("intent","").upper()}</span>
                  <div style="margin-top:4px">{intent_tags or '<span style="color:#3D4F7C;font-size:9px">—</span>'}</div>
                </div>
                <div><div style="font-size:9px;color:#6B7DB3;margin-bottom:4px">Module 2 — Pattern Matching</div>
                  {field_tags} {group_tag} {filter_tags}
                  {'' if any([field_tags, group_tag, filter_tags])
                     else '<span style="color:#3D4F7C;font-size:9px">No entities extracted</span>'}
                </div>
              </div>
            </div>""", unsafe_allow_html=True)

            # Chart
            chart_df = r.get("chart_df", pd.DataFrame())
            if not chart_df.empty:
                st.markdown('<div style="font-size:9px;color:#6B7DB3;font-family:monospace;'
                            'text-transform:uppercase;letter-spacing:.08em;margin:10px 0 4px 0">'
                            "◉ Visualization</div>", unsafe_allow_html=True)
                fig = make_pie_chart(chart_df, r["title"]) if r["chart_type"] == "pie" \
                      else make_bar_chart(chart_df, r["title"])
                if fig:
                    st.pyplot(fig, use_container_width=True)
                    plt.close(fig)

            # Sample table
            sample = r.get("sample_table", pd.DataFrame())
            if not sample.empty:
                st.markdown(f'<div style="font-size:9px;color:#6B7DB3;font-family:monospace;'
                            f'text-transform:uppercase;letter-spacing:.08em;margin:10px 0 4px 0">'
                            f"▤ Filtered Records (first {len(sample)})</div>", unsafe_allow_html=True)
                st.dataframe(sample, use_container_width=True, hide_index=True)

            # Insight
            parts = r["insight"].split("**")
            ins_fmt = ""
            for j, part in enumerate(parts):
                ins_fmt += (f"<strong style='color:#FCD34D'>{part}</strong>"
                            if j % 2 == 1 else part)
            st.markdown(f"""
            <div class="insight-box">
              <div class="insight-title">★ Key Insight</div>
              <div class="insight-text">{ins_fmt}</div>
            </div></div>""", unsafe_allow_html=True)

    # Input form
    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
    default_val = st.session_state.pop("_pending_prompt", "")
    with st.form(key="query_form", clear_on_submit=True):
        col_in, col_btn = st.columns([5, 1])
        with col_in:
            user_prompt = st.text_input(
                "Ask anything about the data", value=default_val,
                placeholder='e.g. "Show average GPA by department"',
                label_visibility="collapsed")
        with col_btn:
            submitted = st.form_submit_button("Run →", use_container_width=True)
    st.markdown('<div style="font-size:9px;color:#3D4F7C;font-family:monospace;margin-top:4px">'
                "Press Run → or Enter · NLP: Intent Detection + Pattern Matching</div>",
                unsafe_allow_html=True)

    if submitted and user_prompt.strip():
        q = user_prompt.strip()
        st.session_state.pipe_step = 3
        st.session_state.messages.append({"role": "user", "text": q})
        with st.spinner("NLP processing → query generation → analysis…"):
            st.session_state.pipe_step = 5
            result = execute_query(st.session_state.df, q)
            st.session_state.pipe_step = 7
        st.session_state.messages.append({"role": "assistant", "result": result})
        st.rerun()


# ─────────────────────────────────────────────
# TAB 2 — ML Risk Model (NEW)
# ─────────────────────────────────────────────
with tab_ml:
    render_ml_risk_panel(st.session_state.df)


# ─────────────────────────────────────────────
# TAB 3 — Advanced Charts (NEW)
# ─────────────────────────────────────────────
with tab_charts:
    render_advanced_charts_panel(st.session_state.df)


# ─────────────────────────────────────────────
# TAB 4 — Export Reports (NEW)
# ─────────────────────────────────────────────
with tab_export:
    st.markdown("""
    <div style="font-size:9px;color:#6B7DB3;font-family:monospace;
    letter-spacing:.08em;text-transform:uppercase;margin-bottom:14px">
    📥 Export Academic Analytics Reports
    </div>
    """, unsafe_allow_html=True)

    col_xl, col_pdf = st.columns(2)

    with col_xl:
        st.markdown("""
        <div style="background:#1A2235;border:1px solid #243050;border-radius:10px;padding:16px">
          <div style="font-size:14px;font-weight:700;color:#EEF2FF;margin-bottom:6px">📊 Excel Report</div>
          <div style="font-size:11px;color:#6B7DB3;line-height:1.6;margin-bottom:12px">
            Generates a styled multi-sheet workbook containing:<br/>
            • <strong style="color:#EEF2FF">Dashboard</strong> — KPIs + GPA by department table<br/>
            • <strong style="color:#EEF2FF">Student Data</strong> — Full color-coded dataset<br/>
            • <strong style="color:#EEF2FF">At-Risk Report</strong> — Highlighted at-risk students<br/>
            • <strong style="color:#EEF2FF">Department Analytics</strong> — Aggregated metrics per dept
          </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        if st.button("⬇ Generate Excel Report", use_container_width=True, key="gen_xl"):
            with st.spinner("Building styled Excel workbook…"):
                xl_bytes = generate_excel_report(st.session_state.df, st.session_state.data_stats)
            st.download_button(
                "📥 Download EduQuery_Report.xlsx",
                xl_bytes,
                file_name="EduQuery_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    with col_pdf:
        st.markdown("""
        <div style="background:#1A2235;border:1px solid #243050;border-radius:10px;padding:16px">
          <div style="font-size:14px;font-weight:700;color:#EEF2FF;margin-bottom:6px">📄 PDF Report</div>
          <div style="font-size:11px;color:#6B7DB3;line-height:1.6;margin-bottom:12px">
            Generates a professional A4 PDF report containing:<br/>
            • <strong style="color:#EEF2FF">KPI Summary</strong> — Key metrics table<br/>
            • <strong style="color:#EEF2FF">GPA Analysis</strong> — Department bar chart<br/>
            • <strong style="color:#EEF2FF">Attendance Distribution</strong> — Banded chart<br/>
            • <strong style="color:#EEF2FF">Performance Bands</strong> — Summary table<br/>
            • <strong style="color:#EEF2FF">At-Risk Report</strong> — Department breakdown<br/>
            • <strong style="color:#EEF2FF">Correlation Analysis</strong> — Pearson r table
          </div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        if st.button("⬇ Generate PDF Report", use_container_width=True, key="gen_pdf"):
            with st.spinner("Rendering PDF report…"):
                pdf_bytes = generate_pdf_report(st.session_state.df, st.session_state.data_stats)
            st.download_button(
                "📥 Download EduQuery_Report.pdf",
                pdf_bytes,
                file_name="EduQuery_Report.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

    st.markdown("---")
    st.markdown("""
    <div style="font-size:9px;color:#6B7DB3;font-family:monospace;
    letter-spacing:.08em;text-transform:uppercase;margin-bottom:8px">
    📋 Quick CSV Export
    </div>
    """, unsafe_allow_html=True)
    col_c1, col_c2, col_c3 = st.columns(3)
    with col_c1:
        csv_all = st.session_state.df.to_csv(index=False)
        st.download_button("⬇ Full Dataset (CSV)", csv_all,
                           file_name="full_dataset.csv", mime="text/csv",
                           use_container_width=True)
    with col_c2:
        at_risk = st.session_state.df[st.session_state.df["Performance_Band"] == "At-Risk"]
        st.download_button("⬇ At-Risk Students (CSV)", at_risk.to_csv(index=False),
                           file_name="at_risk_students.csv", mime="text/csv",
                           use_container_width=True)
    with col_c3:
        dept_sum = (st.session_state.df.groupby("Department")
                    .agg(Students=("Stu_ID","count"), Avg_GPA=("GPA","mean"),
                         Avg_Attendance=("Attendance_Pct","mean"))
                    .round(2).reset_index())
        st.download_button("⬇ Department Summary (CSV)", dept_sum.to_csv(index=False),
                           file_name="dept_summary.csv", mime="text/csv",
                           use_container_width=True)
